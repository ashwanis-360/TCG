import base64
from io import BytesIO

import mysql.connector
import pandas as pd
import json
import base64

import jwt
from fastapi import Depends, HTTPException
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from jwt import PyJWTError
from pydantic import BaseModel
from mysql.connector import Error
from openpyxl.cell import MergedCell
from openpyxl.styles import Side, Font, PatternFill, Alignment, Border
from openpyxl.workbook import Workbook


def get_system_prompt(project_id):
    result = getDBRecord(f"""SELECT * FROM tcg.system_prompt where id=1""")
    result1 = getDBRecord(f"""SELECT * FROM authservice.project where id={project_id}""")
    system_prompt = None
    # Access the value (the first column in the fetched row)
    if result:
        system_prompt = {
            "role": "system",
            "content":
                result['system_prompt'] + "." + result1['description']
        }
    print(system_prompt)
    return system_prompt


def getdb_Connection():
    return mysql.connector.connect(
        host="localhost",  # e.g., "localhost"
        user="root",  # e.g., "root"
        password="admin",  # Your MySQL password
        database="tcg",  # The name of the database
        port=3306

    )


def getDBRecord(query: str, returnAll: bool = False):
    db_connection = mysql.connector.connect(
        host="localhost",  # e.g., "localhost"
        user="root",  # e.g., "root"
        password="admin",  # Your MySQL password
        database="tcg",  # The name of the database
        port=3306

    )
    # Create a cursor object to execute SQL queries
    cursor = db_connection.cursor(dictionary=True)
    project_id = 1
    query = query
    print(query)
    cursor.execute(query)
    result = None
    # Fetch the result (assuming the query returns only one row)
    if returnAll:
        result = cursor.fetchall()
    else:
        result = cursor.fetchone()

    # print("System Prompt read from DB", len(result))
    # Placeholder for storing chat history
    cursor.close()
    db_connection.close()
    return result


def execute_query_with_values(query, values):
    db_connection = None
    try:
        # Establish a connection to the MySQL database
        db_connection = mysql.connector.connect(
            host="localhost",  # e.g., "localhost"
            user="root",  # e.g., "root"
            password="admin",  # Your MySQL password
            database="tcg",  # The name of the database
            port=3306
        )

        if db_connection.is_connected():
            # Create a cursor object to execute SQL queries
            cursor = db_connection.cursor()
            cursor.execute("SET SQL_SAFE_UPDATES = 0;")

            # Execute the query passed as a parameter

            # Execute the query passed as a parameter
            # print(query)
            # print(values)
            cursor.execute(query, values)

            # Commit the transaction
            db_connection.commit()
            new_record_id = cursor.lastrowid

            # print("Query executed successfully.")
            cursor.execute("SET SQL_SAFE_UPDATES = 1;")
            if new_record_id:
                # print(f"New record ID: {new_record_id}")
                return new_record_id

            # print("Query executed successfully.")



    except Error as e:
        print(f"Error while executing the query: {e}")
    finally:
        # Close the cursor and connection
        if db_connection.is_connected():
            cursor.close()
            db_connection.close()


def execute_query(query):
    db_connection = None
    try:
        # Establish a connection to the MySQL database
        db_connection = mysql.connector.connect(
            host="localhost",  # e.g., "localhost"
            user="root",  # e.g., "root"
            password="admin",  # Your MySQL password
            database="tcg",  # The name of the database
            port=3306
        )

        if db_connection.is_connected():
            # Create a cursor object to execute SQL queries
            cursor = db_connection.cursor()
            cursor.execute("SET SQL_SAFE_UPDATES = 0;")

            # Execute the query passed as a parameter

            # Execute the query passed as a parameter
            print(query)
            cursor.execute(query)

            # Commit the transaction
            db_connection.commit()
            new_record_id = cursor.lastrowid

            # print("Query executed successfully.")
            cursor.execute("SET SQL_SAFE_UPDATES = 1;")
            if new_record_id:
                print(f"New record ID: {new_record_id}")
                return new_record_id

            # print("Query executed successfully.")



    except Error as e:
        print(f"Error while executing the query: {e}")
    finally:
        # Close the cursor and connection
        if db_connection.is_connected():
            cursor.close()
            db_connection.close()


def execute_query_param(query, params=None):
    db_connection = None
    try:
        # Establish a connection to the MySQL database
        db_connection = mysql.connector.connect(
            host="localhost",  # e.g., "localhost"
            user="root",  # e.g., "root"
            password="admin",  # Your MySQL password
            database="tcg",  # The name of the database
            port=3306
        )

        if db_connection.is_connected():
            # Create a cursor object to execute SQL queries
            cursor = db_connection.cursor()
            cursor.execute("SET SQL_SAFE_UPDATES = 0;")

            # Execute the query passed as a parameter

            # Execute the query passed as a parameter
            print(query)
            print(params)
            cursor.execute(query, params)

            # Commit the transaction
            db_connection.commit()
            new_record_id = cursor.lastrowid

            # print("Query executed successfully.")
            cursor.execute("SET SQL_SAFE_UPDATES = 1;")
            if new_record_id:
                # print(f"New record ID: {new_record_id}")
                return new_record_id

            # print("Query executed successfully.")



    except Error as e:
        print(f"Error while executing the query: {e}")
    finally:
        # Close the cursor and connection
        if db_connection.is_connected():
            cursor.close()
            db_connection.close()


def extract_json_from_string(complete_string):
    # Find the index of the first '{'
    start_index_curly = complete_string.find('{')
    start_index_square = complete_string.find('[')

    # Determine the first occurrence between '{' or '['
    if start_index_curly == -1:
        start_index = start_index_square
    elif start_index_square == -1:
        start_index = start_index_curly
    else:
        start_index = min(start_index_curly, start_index_square)

    # Find the index of the last '}' or ']'
    end_index_curly = complete_string.rfind('}')
    end_index_square = complete_string.rfind(']')

    # Determine the last occurrence between '}' or ']'
    if end_index_curly == -1:
        end_index = end_index_square
    elif end_index_square == -1:
        end_index = end_index_curly
    else:
        end_index = max(end_index_curly, end_index_square)

    # Check if both start and end index are found
    if start_index != -1 and end_index != -1:
        # Extract the substring containing the JSON object
        json_substring = complete_string[start_index:end_index + 1]
        # print("Actual output",json_substring)
        return json_substring
    else:
        return None  # Return None if no valid JSON object is found


def generate_query_table(data1):
    try:
        data1 = extract_json_from_string(data1)
        df = pd.DataFrame(json.loads(data1))
        html_table = df.to_html(index=False, border=1, justify='center')
    except Exception as e:
        # print(f"*****************n Error occurred in Stage 1: {e} /n*****************")
        html_table = "<p>" + data1.choices[0].message.content + "</p>"

    return html_table


def genrate_asumption_table(data1):
    try:
        data1 = extract_json_from_string(data1)
        df = pd.DataFrame(json.loads(data1))
        html_table = df.to_html(index=False, border=1, justify='center')

    except Exception as e:
        # print(f"*****************/nError Occurred in Stage 2: {e} /n*****************")
        html_table = "<p>" + data1.choices[0].message.content + "</p>"
    return html_table


def generate_userstory_ui(data):
    try:
        # data1 = extract_json_from_string(data)
        html_table = user_story_ui_template(data)
    except Exception as e:
        # print(f"*********/n Error Occurred in Stage 2: {e} /n *******************")
        html_table = "<p>" + data.choices[0].message.content + "</p>"
    return html_table


def generate_requirment_ui(data):
    try:
        # data1 = json.loads(data)
        html_table = requirment_table(data)
    except Exception as e:
        # print(f"*********/n Error Occurred in Stage 4: {e}  /n*********************")
        html_table = "<p>" + data.choices[0].message.content + "</p>"
    return html_table


def generate_requirment_matrix(data):
    try:
        data1 = extract_json_from_string(data)
        json_data = json.loads(data1)
        html_table = generate_html_table_applicability(json_data)
    except Exception as e:
        # print(f"*********/n Error Occurred in Stage 5: {e} /n ********************")
        html_table = "<p>" + data + "</p>"
    return html_table


def user_story_ui_template(json_data):
    if isinstance(json_data, str):
        json_data = json.loads(json_data)
    # Start the HTML table
    html_output = f"""
    <table border="1" cellpadding="5">
        <tr>
            <th>Section</th>
            <th>Details</th>
        </tr>
        <tr>
            <td>Prerequisites</td>
            <td>{json_data['userstory']['prerequisites']}</td>
        </tr>
        <tr>
            <td>User Story Details</td>
            <td>{json_data['userstory']['userstorydetails']}</td>
        </tr>
        <tr>
            <td>Actions</td>
            <td>
                <ul>
    """

    # Add actions as list items
    for action in json_data['userstory']['actions']:
        html_output += f"                    <li>{action}</li>\n"

    html_output += """                </ul>
            </td>
        </tr>
        <tr>
            <td>Test Data Required</td>
            <td>
                <ul>
    """

    # Add test data required as list items
    for testdata in json_data['userstory']['testdatarequired']:
        html_output += f"                    <li>{testdata}</li>\n"

    html_output += """                </ul>
            </td>
        </tr>
        <tr>
            <td>Acceptance Criteria</td>
            <td>
                <ul>
    """

    # Add acceptance criteria as list items
    for criteria in json_data['userstory']['acceptancecriteria']:
        html_output += f"                    <li>{criteria}</li>\n"

    # Close the HTML table
    html_output += """                </ul>
            </td>
        </tr>
    </table>
    """

    return html_output


def testcase_template(data):
    testcaseallrequirment = ""
    try:
        data1 = extract_json_from_string(data)
        testcaseallrequirment = testcaseallrequirment + generate_html_table_testcases(json.loads(data1))
    except Exception as e:
        testcaseallrequirment = testcaseallrequirment + "<p>" + data.choices[0].message.content + "</p>"
    return testcaseallrequirment


def requirment_table(json_data):
    # Start the HTML table

    html = '<table border="1" cellpadding="5" cellspacing="0">'

    # Add table header
    html += '''
    <thead>
        <tr>
            <th>Requirement Detail</th>
            <th>Type</th>
            <th>Test Data</th>
            <th>Steps to Test</th>
        </tr>
    </thead>
    <tbody>
    '''

    # Iterate through the JSON data and add rows to the table
    for requirement in json_data['requirements']:
        html += f'''
        <tr>
            <td>{requirement['requirement_detail']}</td>
            <td>{requirement['type']}</td>
            <td>{requirement['testdata']}</td>
            <td>{requirement['stepstotest']}</td>
        </tr>
        '''

    # Close the table
    html += '</tbody></table>'

    return html


def generate_html_table_applicability(data):
    html = "<table border='1'>"
    html += "<tr><th>Requirement Detail</th><th>Type</th><th>Test Data</th><th>Steps to Test</th><th>Boundary Value Analysis</th><th>Equivalent Class Partitioning</th><th>State Transition Diagram</th><th>Decision Table</th><th>Use Case Testing</th></tr>"

    for req in data['requirements']:
        html += "<tr>"
        html += f"<td>{req['requirement_detail']}</td>"
        html += f"<td>{req['type']}</td>"
        html += f"<td>{req['testdata']}</td>"
        html += f"<td>{req['stepstotest']}</td>"

        # Boundary Value Analysis
        if req['boundary_value_analysis']['applicable']:
            if len(req['boundary_value_analysis']['attributes']) > 0:
                bva_attributes = req['boundary_value_analysis']['attributes']
                bva_details = "<ul>"
                for attr in bva_attributes:
                    bva_details += f"<li>{attr['attribute_name']}: {', '.join(attr['techniquedetails'])}</li>"
                bva_details += "</ul>"
                html += f"<td>{bva_details}</td>"
        else:
            html += "<td>Not Applicable</td>"

        # Equivalent Class Partitioning
        if req['equivalent_class_partitioning']['applicable']:
            if len(req['equivalent_class_partitioning']['attributes']) > 0:
                ecp_attributes = req['equivalent_class_partitioning']['attributes']
                ecp_details = "<ul>"
                for attr in ecp_attributes:
                    ecp_details += f"<li>{attr['attribute_name']}: {', '.join(attr['techniquedetails'])}</li>"
                ecp_details += "</ul>"
                html += f"<td>{ecp_details}</td>"
        else:
            html += "<td>Not Applicable</td>"

        # State Transition Diagram
        if req['state_transition_diagram']['applicable']:
            if len(req['state_transition_diagram']['attributes']) > 0:
                ecp_attributes = req['state_transition_diagram']['attributes']
                ecp_details = "<ul>"
                for attr in ecp_attributes:
                    ecp_details += f"<li>{attr['attribute_name']}: {', '.join(attr['techniquedetails'])}</li>"
                ecp_details += "</ul>"
                html += f"<td>{ecp_details}</td>"
        else:
            html += "<td>Not Applicable</td>"

        # Decision Table
        if req['decision_table']['applicable']:
            if len(req['decision_table']['attributes']) > 0:
                ecp_attributes = req['decision_table']['attributes']
                ecp_details = "<ul>"
                for attr in ecp_attributes:
                    ecp_details += f"<li>{attr['attribute_name']}: {', '.join(attr['techniquedetails'])}</li>"
                ecp_details += "</ul>"
                html += f"<td>{ecp_details}</td>"
        else:
            html += "<td>Not Applicable</td>"

        # Use Case Testing
        if req['use_case_testing']['applicable']:
            if len(req['use_case_testing']['attributes']) > 0:
                ecp_attributes = req['use_case_testing']['attributes']
                ecp_details = "<ul>"
                for attr in ecp_attributes:
                    ecp_details += f"<li>{attr['attribute_name']}: {', '.join(attr['techniquedetails'])}</li>"
                ecp_details += "</ul>"
                html += f"<td>{ecp_details}</td>"
        else:
            html += "<td>Not Applicable</td>"

        html += "</tr>"

    html += "</table>"
    return html


def generate_html_table_testcases(data):
    testcases = data.get('testcases')
    if not testcases:
        html = f"""
            <html>
            <head>
                <title>Test Cases</title>
                <style>
                    table, th, td {{
                        border: 1px solid black;
                        border-collapse: collapse;
                        padding: 8px;
                    }}
                    th {{
                        background-color: #f2f2f2;
                    }}
                </style>
            </head>
            <body>
                <h2>Test Cases</h2>
                <p><strong>Requirement:</strong> {data['requirement']}</p>
                <p><strong>Techniques:</strong> {data['techniques']}</p>
                <table>
                    <tr>
                        <th>Test Case Summary</th>
                        <th>Test Steps</th>
                        <th>Expected Result</th>
                        <th>Test Data</th>
                    </tr>
            """

        html += """
                </table>
            </body>
            </html>
            """
        return html
    html = f"""
    <html>
    <head>
        <title>Test Cases</title>
        <style>
            table, th, td {{
                border: 1px solid black;
                border-collapse: collapse;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
        </style>
    </head>
    <body>
        <h2>Test Cases</h2>
        <p><strong>Requirement:</strong> {data['requirement']}</p>
        <p><strong>Techniques:</strong> {data['techniques']}</p>
        <table>
            <tr>
                <th>Test Case Summary</th>
                <th>Test Steps</th>
                <th>Expected Result</th>
                <th>Test Data</th>
            </tr>
    """

    for testcase in data['testcases']:
        html += f"""
            <tr>
                <td>{testcase['testcase_summary']}</td>
                <td>{testcase['test_steps']}</td>
                <td>{testcase['expected_result']}</td>
                <td>{testcase['test_data']}</td>
            </tr>
        """

    html += """
        </table>
    </body>
    </html>
    """

    return html


def safe_json_load(value):
    if isinstance(value, (list, dict)):
        return value
    if isinstance(value, bytes):
        value = value.decode('utf-8')
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            print("Failed to parse JSON:", value)
            return []
    return []


def fetch_all(query, params=None):
    connection = None
    try:
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="admin",
            database="tcg",
            port=3306
        )
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query, params)
        results = cursor.fetchall()
        return results
    except mysql.connector.Error as e:
        print(f"MySQL Error: {e}")
        return []
    finally:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()


# def prepare_excel(pr_id, user_story_id, df):
#     # Handle test_steps and other preprocessing
#     df['test_steps'] = df['test_steps'].apply(eval)
#     exploded_df = df.explode('test_steps')
#
#     # Create an Excel workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Test Cases"
#
#     # Add a title row above the header row
#     title = f"Test Cases for PR: {pr_id}, User Story: {user_story_id}"
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(exploded_df.columns))
#     title_cell = ws.cell(row=1, column=1, value=title)
#     title_cell.font = Font(bold=True, size=14, color="FFFFFF")
#     title_cell.alignment = Alignment(horizontal="center", vertical="center")
#     title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Dark blue fill
#
#     # Add headers with formatting
#     header_row = list(exploded_df.columns)
#     header_row = [header.replace('_', ' ').upper() for header in header_row]  # Capitalize header
#     light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
#     white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
#
#     for c_idx, header in enumerate(header_row, start=1):
#         cell = ws.cell(row=2, column=c_idx, value=header)
#         cell.alignment = Alignment(wrap_text=True)
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background
#     border_style = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )
#     # Write data
#     current_row = 3  # Start after the header row
#     for row in dataframe_to_rows(exploded_df, index=False, header=False):
#         for c_idx, value in enumerate(row, start=1):
#             cell = ws.cell(row=current_row, column=c_idx, value=value)
#             cell.alignment = Alignment(wrap_text=True)
#             cell.border = border_style
#             if current_row % 2 == 0:
#                 cell.fill = light_blue_fill
#             else:
#                 cell.fill = white_fill
#
#         # Add a blank row after each test case with gray fill
#
#         current_row += 1  # Move to the next data row (skip blank row)
#
#     # Adjust column widths to fit the content
#     for col in ws.columns:
#         max_length = 0
#         column = col[0].coordinate[0]
#         # Get the column name
#         for cell in col:
#             try:
#                 if cell.row == 0 or cell.row == 1:  # Check if the row index is 1 (the first row)
#                     continue
#                 # Skip merged cells
#                 if cell and not (isinstance(cell, MergedCell)):
#
#                     if cell.value and len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#             except Exception as e:
#                 print(f"Error processing cell {cell}: {e}")
#                 pass
#         adjusted_width = (50 if max_length > 50 else max_length + 2)
#         ws.column_dimensions[column].width = adjusted_width
#
#     # Save to memory
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#     return output


def prepare_excel2(pr_id, user_story_id, df):
    # Handle test_steps and other preprocessing
    df['test_steps'] = df['test_steps'].apply(eval)
    df['test_data'] = df['test_data'].apply(eval)

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Add a title row above the header row
    title = f"Test Cases for PR: {pr_id}, User Story: {user_story_id}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Dark blue fill

    # Add headers with formatting
    header_row = list(df.columns)
    header_row = [header.replace('_', ' ').upper() for header in header_row]  # Capitalize header
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    for c_idx, header in enumerate(header_row, start=1):
        cell = ws.cell(row=2, column=c_idx, value=header)
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue background

    # Set the border style for cells
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Write data: Exploding test_steps into separate rows while keeping other columns fixed
    current_row = 3  # Start after the header row
    for idx, row in df.iterrows():
        # Find the longest list in the row (for test_steps or any other list-type columns)
        # max_len = max([len(row[col]) if isinstance(row[col], list) else 1 for col in df.columns])
        #
        # # Iterate over the number of rows needed (based on the longest list length)
        # for i in range(max_len):
        for c_idx, column in enumerate(df.columns, start=1):
            # If the column contains a list, pick the ith item from the list, or the first item if the list is shorter
            value = row[column]

            # If the value is a list, join the items with a newline
            if isinstance(value, list):
                value = "\n".join(str(item) for item in value)
            cell = ws.cell(row=current_row, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border_style

            # Alternating row colors
            if current_row % 2 == 0:
                cell.fill = light_blue_fill
            else:
                cell.fill = white_fill

        current_row += 1  # Move to the next row for the next set of data

    # Adjust column widths to fit the content
    for col in ws.columns:
        max_length = 0
        column = col[0].coordinate[0]
        for cell in col:
            try:
                if cell.row == 0 or cell.row == 1:  # Skip header row
                    continue
                if cell and not isinstance(cell, MergedCell):
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error processing cell {cell}: {e}")
                pass
        adjusted_width = (50 if max_length > 50 else max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def encode_data(chat_history):
    encoded_chat_history = base64.b64encode(json.dumps(chat_history).encode("utf-8"))
    return encoded_chat_history


def before_comm_logging(chat_history, pr_id, userstory_ref, stage):
    log_id = execute_query_param(
        f"""INSERT INTO `tcg`.`com_log`(`project_id`, `userstory_id`, `stage`, `before_chat`, `after_chat`) 
                VALUES({pr_id}, {userstory_ref}, {stage}, "{encode_data(chat_history)}", null);"""
    )
    return log_id


def after_comm_logging(log_id, chat_history):
    execute_query_param(
        f"""UPDATE `tcg`.`com_log` SET `after_chat` = "{encode_data(chat_history)}" where id = {log_id};""")


def log_stage_output(pr_id, user_story_ref, stage, json_structure_data):
    latest_response = execute_query_param(
        f"""INSERT INTO `tcg`.`stage_output`(`project_id`,`userstory_id`,`stage`,`output`) 
           VALUES({pr_id}, {user_story_ref}, {stage}, "{encode_data(json_structure_data)}");"""
    )
    return latest_response


SECRET_KEY = "mySecretKeymySecretKeymySecretKeymySecretKey"  # Replace this with your Spring Boot secret key
ALGORITHM = "HS256"  # Algorithm used in your Spring Boot JWT generation

security = HTTPBearer()


class TokenData(BaseModel):
    username: str
    roles: list = []
    token: str


def verify_jwt(token: str):
    try:
        # Decode the JWT token using the secret key and algorithm
        secret = base64.b64decode(SECRET_KEY)

        payload = jwt.decode(token, secret, algorithms=["HS256"])
        username: str = payload.get("sub")
        roles: list = payload.get("roles", [])
        btoken: str = token
        print(username, roles, btoken)
        if username is None:
            raise jwt.InvalidTokenError
        return TokenData(username=username, roles=[roles], token=btoken)
    except PyJWTError:
        raise HTTPException(status_code=401, detail="Token is invalid or expired")


# This will be used to extract the token from the request
def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    token = credentials.credentials
    print(token)
    return verify_jwt(token)


def create_bold_paragraph(text):
    return {
        "type": "paragraph",
        "content": [
            {
                "type": "text",
                "text": text,
                "marks": [{"type": "strong"}]
            }
        ]
    }

def create_ordered_list(items):
    return {
        "type": "orderedList",
        "content": [
            {
                "type": "listItem",
                "content": [
                    {
                        "type": "paragraph",
                        "content": [{"type": "text", "text": item}]
                    }
                ]
            } for item in items
        ]
    }
def extract_text_from_adf(node):
    """
    Recursively extracts text from Jira's Atlassian Document Format JSON.
    """
    if not isinstance(node, dict):
        return ""
    node_type = node.get("type")
    result = ""

    if node_type == "text":
        result += node.get("text", "")

    # Recurse into content arrays
    if "content" in node:
        for child in node["content"]:
            result += extract_text_from_adf(child)
        if node_type == "paragraph":
            result += "\n"  # add line breaks between paragraphs

    return result.strip()
